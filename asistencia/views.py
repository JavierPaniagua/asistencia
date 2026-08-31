# ============================================================
# ASISTENCIA - VIEWS.PY
# ============================================================

import os
import calendar

from io import BytesIO
from datetime import datetime, date, timedelta

from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone


# ============================================================
# EXCEL
# ============================================================

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter


# ============================================================
# PDF
# ============================================================

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape
from reportlab.lib.styles import (
    getSampleStyleSheet,
    ParagraphStyle,
)
from reportlab.lib.enums import (
    TA_CENTER,
    TA_LEFT,
)
from reportlab.lib.units import cm, inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)


# ============================================================
# MODELOS
# ============================================================

from .models import (
    Alumno,
    Asistencia,
    Materia,
    HorarioClase,
    ClaseActiva,
)


# ============================================================
# HOJA OFICIO
# 8,5 x 13 pulgadas
# ============================================================

OFICIO = (
    8.5 * inch,
    13 * inch,
)


# ============================================================
# ESTADO GLOBAL DE PANTALLA
# ============================================================

ultimo_evento = {
    "evento_id": 0,
    "estado": "esperando",
}


# ============================================================
# ACTUALIZAR EVENTO
# ============================================================

def actualizar_evento(datos):
    global ultimo_evento

    nuevo_id = ultimo_evento.get(
        "evento_id",
        0
    ) + 1

    ultimo_evento = {
        "evento_id": nuevo_id,
        **datos,
    }

    return ultimo_evento


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def nombre_alumno(alumno):
    return (
        f"{alumno.apellidos}, "
        f"{alumno.nombres}"
    ).strip()


def curso_alumno(alumno):
    return alumno.curso or ""


def cedula_alumno(alumno):
    return alumno.cedula or ""


# ============================================================
# OBTENER HORARIO ACTUAL
# ============================================================

def obtener_horario_actual():

    ahora = timezone.localtime()

    fecha_actual = ahora.date()
    hora_actual = ahora.time()

    return (
        HorarioClase.objects
        .filter(
            anio_lectivo=fecha_actual.year,
            dia_semana=fecha_actual.weekday(),
            hora_inicio__lte=hora_actual,
            hora_fin__gt=hora_actual,
            activo=True,
            materia__activo=True,
        )
        .select_related("materia")
        .order_by("hora_inicio")
        .first()
    )


# ============================================================
# OBTENER CLASE ACTUAL
# ============================================================

def obtener_clase_actual():

    horario = obtener_horario_actual()

    # --------------------------------------------------------
    # HORARIO AUTOMÁTICO
    # --------------------------------------------------------

    if horario:

        return {
            "tipo": "horario",
            "materia": horario.materia,
            "curso": horario.curso,
            "hora_inicio": horario.hora_inicio,
            "hora_fin": horario.hora_fin,
            "horario": horario,
        }

    # --------------------------------------------------------
    # CLASE MANUAL
    # --------------------------------------------------------

    clase_manual = (
        ClaseActiva.objects
        .filter(
            activa=True,
            materia__activo=True,
        )
        .select_related("materia")
        .order_by("-iniciada")
        .first()
    )

    if clase_manual:

        return {
            "tipo": "manual",
            "materia": clase_manual.materia,
            "curso": None,
            "hora_inicio": None,
            "hora_fin": None,
            "clase_manual": clase_manual,
        }

    return None


# ============================================================
# PANTALLA
# ============================================================

def pantalla(request):

    return render(
        request,
        "asistencia/pantalla.html",
    )


# ============================================================
# ESTADO PANTALLA
# ============================================================

def estado_pantalla(request):

    return JsonResponse(
        ultimo_evento,
        json_dumps_params={
            "ensure_ascii": False
        },
    )


# ============================================================
# ESTADO CLASE
# ============================================================

def estado_clase(request):

    clase = obtener_clase_actual()

    if not clase:

        return JsonResponse(
            {
                "activa": False,
                "estado": "sin_clase",
                "mensaje": (
                    "SIN CLASE PROGRAMADA"
                ),
            },
            json_dumps_params={
                "ensure_ascii": False
            },
        )

    respuesta = {
        "activa": True,
        "estado": "clase_activa",
        "materia": clase[
            "materia"
        ].nombre,
        "curso": clase.get(
            "curso"
        ) or "",
        "tipo": clase[
            "tipo"
        ],
    }

    if clase.get("hora_inicio"):

        respuesta["hora_inicio"] = (
            clase["hora_inicio"]
            .strftime("%H:%M")
        )

    if clase.get("hora_fin"):

        respuesta["hora_fin"] = (
            clase["hora_fin"]
            .strftime("%H:%M")
        )

    return JsonResponse(
        respuesta,
        json_dumps_params={
            "ensure_ascii": False
        },
    )


# ============================================================
# REGISTRAR ALUMNO
# ============================================================

def registrar_alumno(alumno):

    clase = obtener_clase_actual()

    if not clase:

        return actualizar_evento(
            {
                "estado": "sin_clase",
                "nombre": nombre_alumno(
                    alumno
                ),
                "curso": curso_alumno(
                    alumno
                ),
                "cedula": cedula_alumno(
                    alumno
                ),
                "mensaje": (
                    "No hay una clase "
                    "programada en este momento."
                ),
            }
        )

    materia = clase["materia"]
    curso_clase = clase.get("curso")

    # --------------------------------------------------------
    # VERIFICAR CURSO
    # --------------------------------------------------------

    if (
        curso_clase
        and alumno.curso != curso_clase
    ):

        return actualizar_evento(
            {
                "estado": "error",
                "nombre": nombre_alumno(
                    alumno
                ),
                "curso": curso_alumno(
                    alumno
                ),
                "cedula": cedula_alumno(
                    alumno
                ),
                "materia": materia.nombre,
                "mensaje": (
                    f"El alumno pertenece a "
                    f"{alumno.curso} y la clase "
                    f"corresponde a "
                    f"{curso_clase}."
                ),
            }
        )

    fecha_actual = timezone.localdate()

    ahora = timezone.localtime()

    hora_actual = (
        ahora.time()
        .replace(
            microsecond=0
        )
    )

    # --------------------------------------------------------
    # DUPLICADO
    # --------------------------------------------------------

    asistencia_existente = (
        Asistencia.objects
        .filter(
            alumno=alumno,
            materia=materia,
            fecha=fecha_actual,
        )
        .order_by("hora")
        .first()
    )

    if asistencia_existente:

        hora_existente = ""

        if asistencia_existente.hora:

            hora_existente = (
                asistencia_existente
                .hora
                .strftime("%H:%M:%S")
            )

        return actualizar_evento(
            {
                "estado": "duplicado",
                "nombre": nombre_alumno(
                    alumno
                ),
                "curso": curso_alumno(
                    alumno
                ),
                "cedula": cedula_alumno(
                    alumno
                ),
                "materia": materia.nombre,
                "fecha": (
                    fecha_actual
                    .strftime("%d/%m/%Y")
                ),
                "hora": hora_existente,
                "mensaje": (
                    "La asistencia ya fue "
                    "registrada para esta materia."
                ),
            }
        )

    # --------------------------------------------------------
    # CREAR
    # --------------------------------------------------------

    asistencia = (
        Asistencia.objects
        .create(
            alumno=alumno,
            materia=materia,
            fecha=fecha_actual,
            hora=hora_actual,
        )
    )

    return actualizar_evento(
        {
            "estado": "registrado",
            "nombre": nombre_alumno(
                alumno
            ),
            "curso": curso_alumno(
                alumno
            ),
            "cedula": cedula_alumno(
                alumno
            ),
            "materia": materia.nombre,
            "fecha": (
                fecha_actual
                .strftime("%d/%m/%Y")
            ),
            "hora": (
                hora_actual
                .strftime("%H:%M:%S")
            ),
            "mensaje": (
                "Asistencia registrada "
                "correctamente"
            ),
            "asistencia_id": (
                asistencia.id
            ),
        }
    )


# ============================================================
# REGISTRO RFID
# ============================================================

def registrar_rfid(request):

    uid = request.GET.get(
        "uid",
        ""
    )

    uid = (
        uid
        .replace(" ", "")
        .replace(":", "")
        .replace("-", "")
        .upper()
        .strip()
    )

    if not uid:

        return JsonResponse(
            {
                "estado": "error",
                "mensaje": (
                    "UID RFID no recibido"
                ),
            },
            status=400,
            json_dumps_params={
                "ensure_ascii": False
            },
        )

    try:

        alumno = Alumno.objects.get(
            uid_rfid=uid,
            activo=True,
        )

    except Alumno.DoesNotExist:

        evento = actualizar_evento(
            {
                "estado": "desconocido",
                "uid": uid,
                "mensaje": (
                    "Tarjeta RFID "
                    "no registrada"
                ),
            }
        )

        return JsonResponse(
            evento,
            json_dumps_params={
                "ensure_ascii": False
            },
        )

    evento = registrar_alumno(
        alumno
    )

    return JsonResponse(
        evento,
        json_dumps_params={
            "ensure_ascii": False
        },
    )


# ============================================================
# REGISTRO POR CÉDULA
# ============================================================

def registrar_por_cedula(request):

    cedula = (
        request.GET.get(
            "cedula",
            ""
        )
        .strip()
    )

    if not cedula:

        return JsonResponse(
            {
                "estado": "error",
                "mensaje": (
                    "Ingrese el número "
                    "de cédula"
                ),
            },
            status=400,
            json_dumps_params={
                "ensure_ascii": False
            },
        )

    if not cedula.isdigit():

        return JsonResponse(
            {
                "estado": "error",
                "mensaje": (
                    "La cédula debe contener "
                    "solamente números"
                ),
            },
            status=400,
            json_dumps_params={
                "ensure_ascii": False
            },
        )

    try:

        alumno = Alumno.objects.get(
            cedula=cedula,
            activo=True,
        )

    except Alumno.DoesNotExist:

        evento = actualizar_evento(
            {
                "estado": "desconocido",
                "cedula": cedula,
                "mensaje": (
                    "Cédula no registrada"
                ),
            }
        )

        return JsonResponse(
            evento,
            json_dumps_params={
                "ensure_ascii": False
            },
        )

    evento = registrar_alumno(
        alumno
    )

    return JsonResponse(
        evento,
        json_dumps_params={
            "ensure_ascii": False
        },
    )


# ============================================================
# PRUEBA
# ============================================================

def probar_asistencia(
    request,
    cedula
):

    try:

        alumno = Alumno.objects.get(
            cedula=cedula,
            activo=True,
        )

    except Alumno.DoesNotExist:

        return JsonResponse(
            {
                "estado": "error",
                "mensaje": (
                    "Alumno no encontrado"
                ),
            },
            status=404,
            json_dumps_params={
                "ensure_ascii": False
            },
        )

    evento = registrar_alumno(
        alumno
    )

    return JsonResponse(
        evento,
        json_dumps_params={
            "ensure_ascii": False
        },
    )


# ============================================================
# INICIAR CLASE MANUAL
# ============================================================

def iniciar_clase(request):

    materia_id = request.GET.get(
        "materia"
    )

    if not materia_id:

        return JsonResponse(
            {
                "estado": "error",
                "mensaje": (
                    "Debe indicar la materia"
                ),
            },
            status=400,
        )

    materia = (
        Materia.objects
        .filter(
            id=materia_id,
            activo=True,
        )
        .first()
    )

    if not materia:

        return JsonResponse(
            {
                "estado": "error",
                "mensaje": (
                    "Materia no encontrada"
                ),
            },
            status=404,
        )

    clases_anteriores = (
        ClaseActiva.objects
        .filter(
            activa=True
        )
    )

    for clase in clases_anteriores:

        clase.activa = False

        clase.finalizada = (
            timezone.now()
        )

        clase.save(
            update_fields=[
                "activa",
                "finalizada",
            ]
        )

    clase = (
        ClaseActiva.objects
        .create(
            materia=materia,
            activa=True,
        )
    )

    return JsonResponse(
        {
            "estado": "ok",
            "mensaje": (
                "Clase iniciada"
            ),
            "materia": (
                materia.nombre
            ),
            "clase_id": (
                clase.id
            ),
        },
        json_dumps_params={
            "ensure_ascii": False
        },
    )


# ============================================================
# FINALIZAR CLASE
# ============================================================

def finalizar_clase(request):

    clase = (
        ClaseActiva.objects
        .filter(
            activa=True
        )
        .order_by("-iniciada")
        .first()
    )

    if not clase:

        return JsonResponse(
            {
                "estado": "sin_clase",
                "mensaje": (
                    "No existe una "
                    "clase manual activa"
                ),
            }
        )

    clase.activa = False

    clase.finalizada = (
        timezone.now()
    )

    clase.save(
        update_fields=[
            "activa",
            "finalizada",
        ]
    )

    return JsonResponse(
        {
            "estado": "ok",
            "mensaje": (
                "Clase finalizada"
            ),
            "materia": (
                clase.materia.nombre
            ),
        },
        json_dumps_params={
            "ensure_ascii": False
        },
    )


# ============================================================
# REPORTES
# ============================================================

def reportes_inicio(request):

    return render(
        request,
        "asistencia/reportes_inicio.html",
    )


# ============================================================
# REPORTE POR FECHA
# ============================================================

def reporte_asistencia(request):

    hoy = timezone.localdate()
    ahora = timezone.localtime()

    fecha_str = request.GET.get(
        "fecha"
    )

    curso = request.GET.get(
        "curso",
        "1° BTI"
    )

    materia_id = request.GET.get(
        "materia"
    )

    if fecha_str:

        try:

            fecha = datetime.strptime(
                fecha_str,
                "%Y-%m-%d"
            ).date()

        except ValueError:

            fecha = hoy

    else:

        fecha = hoy

    cursos = (
        Alumno.objects
        .filter(
            activo=True
        )
        .values_list(
            "curso",
            flat=True
        )
        .distinct()
        .order_by("curso")
    )

    materias = (
        Materia.objects
        .filter(
            activo=True
        )
        .order_by("nombre")
    )

    materia = None

    if materia_id:

        materia = (
            Materia.objects
            .filter(
                id=materia_id,
                activo=True
            )
            .first()
        )

    if materia is None:

        materia = materias.first()

    alumnos = (
        Alumno.objects
        .filter(
            activo=True,
            curso=curso,
        )
        .order_by(
            "apellidos",
            "nombres"
        )
    )

    horario = None

    if materia:

        horario = (
            HorarioClase.objects
            .filter(
                anio_lectivo=fecha.year,
                curso=curso,
                materia=materia,
                dia_semana=(
                    fecha.weekday()
                ),
                activo=True,
            )
            .order_by(
                "hora_inicio"
            )
            .first()
        )

    asistencias = (
        Asistencia.objects
        .filter(
            fecha=fecha,
            alumno__curso=curso,
            materia=materia,
        )
        .select_related(
            "alumno",
            "materia"
        )
    )

    mapa = {}

    for asistencia in asistencias:

        if (
            asistencia.alumno_id
            not in mapa
        ):

            mapa[
                asistencia.alumno_id
            ] = asistencia

    clase_finalizada = False

    if horario:

        if fecha < hoy:

            clase_finalizada = True

        elif fecha > hoy:

            clase_finalizada = False

        else:

            clase_finalizada = (
                ahora.time()
                >= horario.hora_fin
            )

    filas = []

    presentes = 0
    ausentes = 0
    pendientes = 0

    for numero, alumno in enumerate(
        alumnos,
        start=1
    ):

        asistencia = mapa.get(
            alumno.id
        )

        if asistencia:

            estado = "PRESENTE"

            hora = (
                asistencia.hora.strftime(
                    "%H:%M"
                )
                if asistencia.hora
                else ""
            )

            presentes += 1

        else:

            hora = ""

            if not horario:

                estado = "SIN CLASE"

            elif clase_finalizada:

                estado = "AUSENTE"

                ausentes += 1

            else:

                estado = "PENDIENTE"

                pendientes += 1

        filas.append(
            {
                "numero": numero,
                "alumno": alumno,
                "hora": hora,
                "estado": estado,
            }
        )

    total_alumnos = alumnos.count()

    clases_realizadas = (
        presentes
        + ausentes
    )

    porcentaje = (
        round(
            presentes
            * 100
            / clases_realizadas,
            1
        )
        if clases_realizadas
        else 0
    )

    return render(
        request,
        "asistencia/reporte.html",
        {
            "fecha": fecha,
            "curso": curso,
            "cursos": cursos,
            "materia": materia,
            "materias": materias,
            "horario": horario,
            "filas": filas,
            "total_alumnos": (
                total_alumnos
            ),
            "presentes": presentes,
            "ausentes": ausentes,
            "pendientes": pendientes,
            "porcentaje": porcentaje,
        },
    )


# ============================================================
# REPORTE MENSUAL
# ============================================================

def reporte_mensual_curso(request):

    hoy = timezone.localdate()
    ahora = timezone.localtime()

    curso = request.GET.get(
        "curso",
        "1° BTI"
    )

    materia_id = request.GET.get(
        "materia"
    )

    try:

        mes = int(
            request.GET.get(
                "mes",
                hoy.month
            )
        )

    except (
        TypeError,
        ValueError
    ):

        mes = hoy.month

    try:

        anio = int(
            request.GET.get(
                "anio",
                hoy.year
            )
        )

    except (
        TypeError,
        ValueError
    ):

        anio = hoy.year

    cursos = (
        Alumno.objects
        .filter(
            activo=True
        )
        .values_list(
            "curso",
            flat=True
        )
        .distinct()
        .order_by("curso")
    )

    materias = (
        Materia.objects
        .filter(
            activo=True
        )
        .order_by("nombre")
    )

    materia = None

    if materia_id:

        materia = (
            Materia.objects
            .filter(
                id=materia_id,
                activo=True
            )
            .first()
        )

    if materia is None:

        materia = materias.first()

    if not materia:

        return HttpResponse(
            "No existe una materia activa.",
            status=400
        )

    primer_dia = date(
        anio,
        mes,
        1
    )

    ultimo_numero = (
        calendar.monthrange(
            anio,
            mes
        )[1]
    )

    ultimo_dia = date(
        anio,
        mes,
        ultimo_numero
    )

    horarios = (
        HorarioClase.objects
        .filter(
            anio_lectivo=anio,
            curso=curso,
            materia=materia,
            activo=True,
        )
        .order_by(
            "dia_semana",
            "hora_inicio"
        )
    )

    sesiones = []

    fecha_actual = primer_dia

    while fecha_actual <= ultimo_dia:

        horarios_dia = (
            horarios.filter(
                dia_semana=(
                    fecha_actual.weekday()
                )
            )
        )

        for horario in horarios_dia:

            if fecha_actual < hoy:

                finalizada = True

            elif fecha_actual > hoy:

                finalizada = False

            else:

                finalizada = (
                    ahora.time()
                    >= horario.hora_fin
                )

            sesiones.append(
                {
                    "fecha": fecha_actual,
                    "horario": horario,
                    "finalizada": (
                        finalizada
                    ),
                }
            )

        fecha_actual += timedelta(
            days=1
        )

    alumnos = (
        Alumno.objects
        .filter(
            activo=True,
            curso=curso
        )
        .order_by(
            "apellidos",
            "nombres"
        )
    )

    asistencias = (
        Asistencia.objects
        .filter(
            alumno__curso=curso,
            materia=materia,
            fecha__range=[
                primer_dia,
                ultimo_dia
            ],
        )
        .select_related(
            "alumno",
            "materia"
        )
    )

    mapa = {}

    for asistencia in asistencias:

        clave = (
            asistencia.alumno_id,
            asistencia.fecha
        )

        if clave not in mapa:

            mapa[clave] = asistencia

    filas = []

    for numero, alumno in enumerate(
        alumnos,
        start=1
    ):

        celdas = []

        presentes = 0
        ausentes = 0
        pendientes = 0

        for sesion in sesiones:

            asistencia = mapa.get(
                (
                    alumno.id,
                    sesion["fecha"]
                )
            )

            if asistencia:

                estado = "PRESENTE"

                hora = (
                    asistencia.hora.strftime(
                        "%H:%M"
                    )
                    if asistencia.hora
                    else ""
                )

                presentes += 1

            else:

                hora = ""

                if sesion["finalizada"]:

                    estado = "AUSENTE"

                    ausentes += 1

                else:

                    estado = "PENDIENTE"

                    pendientes += 1

            celdas.append(
                {
                    "fecha": (
                        sesion["fecha"]
                    ),
                    "estado": estado,
                    "hora": hora,
                }
            )

        realizadas = (
            presentes
            + ausentes
        )

        porcentaje = (
            round(
                presentes
                * 100
                / realizadas,
                1
            )
            if realizadas
            else 0
        )

        filas.append(
            {
                "numero": numero,
                "alumno": alumno,
                "celdas": celdas,
                "presentes": presentes,
                "ausentes": ausentes,
                "pendientes": pendientes,
                "porcentaje": porcentaje,
            }
        )

    meses = [
        (1, "ENERO"),
        (2, "FEBRERO"),
        (3, "MARZO"),
        (4, "ABRIL"),
        (5, "MAYO"),
        (6, "JUNIO"),
        (7, "JULIO"),
        (8, "AGOSTO"),
        (9, "SEPTIEMBRE"),
        (10, "OCTUBRE"),
        (11, "NOVIEMBRE"),
        (12, "DICIEMBRE"),
    ]

    return render(
        request,
        (
            "asistencia/"
            "reporte_mensual_curso.html"
        ),
        {
            "curso": curso,
            "cursos": cursos,
            "materia": materia,
            "materias": materias,
            "mes": mes,
            "meses": meses,
            "nombre_mes": dict(
                meses
            ).get(
                mes,
                ""
            ),
            "anio": anio,
            "sesiones": sesiones,
            "filas": filas,
            "total_clases": len(
                sesiones
            ),
        },
    )


# ============================================================
# REPORTE INDIVIDUAL
# ============================================================

def reporte_individual(request):

    hoy = timezone.localdate()
    ahora = timezone.localtime()

    alumnos = (
        Alumno.objects
        .filter(
            activo=True
        )
        .order_by(
            "apellidos",
            "nombres"
        )
    )

    alumno_id = request.GET.get(
        "alumno"
    )

    tipo = request.GET.get(
        "tipo",
        "mensual"
    )

    alumno = None

    if alumno_id:

        alumno = (
            Alumno.objects
            .filter(
                id=alumno_id,
                activo=True
            )
            .first()
        )

    try:

        mes = int(
            request.GET.get(
                "mes",
                hoy.month
            )
        )

    except (
        TypeError,
        ValueError
    ):

        mes = hoy.month

    try:

        anio = int(
            request.GET.get(
                "anio",
                hoy.year
            )
        )

    except (
        TypeError,
        ValueError
    ):

        anio = hoy.year

    fecha_semana = request.GET.get(
        "fecha"
    )

    if fecha_semana:

        try:

            fecha_base = datetime.strptime(
                fecha_semana,
                "%Y-%m-%d"
            ).date()

        except ValueError:

            fecha_base = hoy

    else:

        fecha_base = hoy

    inicio_semana = (
        fecha_base
        - timedelta(
            days=fecha_base.weekday()
        )
    )

    fin_semana = (
        inicio_semana
        + timedelta(days=6)
    )

    if tipo == "semanal":

        fecha_inicio = inicio_semana
        fecha_fin = fin_semana

    else:

        fecha_inicio = date(
            anio,
            mes,
            1
        )

        fecha_fin = date(
            anio,
            mes,
            calendar.monthrange(
                anio,
                mes
            )[1]
        )

    filas = []

    presentes = 0
    ausentes = 0
    pendientes = 0

    if alumno:

        horarios = (
            HorarioClase.objects
            .filter(
                curso=alumno.curso,
                activo=True,
                materia__activo=True
            )
            .select_related("materia")
            .order_by(
                "dia_semana",
                "hora_inicio"
            )
        )

        fecha_actual = fecha_inicio

        while fecha_actual <= fecha_fin:

            horarios_dia = (
                horarios.filter(
                    dia_semana=(
                        fecha_actual.weekday()
                    ),
                    anio_lectivo=(
                        fecha_actual.year
                    )
                )
            )

            for horario in horarios_dia:

                asistencia = (
                    Asistencia.objects
                    .filter(
                        alumno=alumno,
                        materia=horario.materia,
                        fecha=fecha_actual,
                    )
                    .order_by("hora")
                    .first()
                )

                if asistencia:

                    estado = "PRESENTE"

                    hora_registro = (
                        asistencia.hora
                        .strftime("%H:%M")
                        if asistencia.hora
                        else ""
                    )

                    presentes += 1

                else:

                    hora_registro = ""

                    if fecha_actual < hoy:

                        estado = "AUSENTE"

                        ausentes += 1

                    elif fecha_actual > hoy:

                        estado = "PENDIENTE"

                        pendientes += 1

                    elif (
                        ahora.time()
                        >= horario.hora_fin
                    ):

                        estado = "AUSENTE"

                        ausentes += 1

                    else:

                        estado = "PENDIENTE"

                        pendientes += 1

                filas.append(
                    {
                        "fecha": fecha_actual,
                        "materia": (
                            horario.materia
                        ),
                        "hora_inicio": (
                            horario.hora_inicio
                        ),
                        "hora_fin": (
                            horario.hora_fin
                        ),
                        "hora_registro": (
                            hora_registro
                        ),
                        "estado": estado,
                    }
                )

            fecha_actual += timedelta(
                days=1
            )

    filas.sort(
        key=lambda fila: (
            fila["fecha"],
            fila["hora_inicio"],
        )
    )

    realizadas = (
        presentes
        + ausentes
    )

    porcentaje = (
        round(
            presentes
            * 100
            / realizadas,
            1
        )
        if realizadas
        else 0
    )

    meses = [
        (1, "ENERO"),
        (2, "FEBRERO"),
        (3, "MARZO"),
        (4, "ABRIL"),
        (5, "MAYO"),
        (6, "JUNIO"),
        (7, "JULIO"),
        (8, "AGOSTO"),
        (9, "SEPTIEMBRE"),
        (10, "OCTUBRE"),
        (11, "NOVIEMBRE"),
        (12, "DICIEMBRE"),
    ]

    return render(
        request,
        (
            "asistencia/"
            "reporte_individual.html"
        ),
        {
            "alumnos": alumnos,
            "alumno": alumno,
            "tipo": tipo,
            "mes": mes,
            "meses": meses,
            "nombre_mes": dict(
                meses
            ).get(
                mes,
                ""
            ),
            "anio": anio,
            "fecha_base": fecha_base,
            "inicio_semana": (
                inicio_semana
            ),
            "fin_semana": fin_semana,
            "filas": filas,
            "presentes": presentes,
            "ausentes": ausentes,
            "pendientes": pendientes,
            "porcentaje": porcentaje,
            "total_clases": len(
                filas
            ),
            "clases_realizadas": (
                realizadas
            ),
        },
    )


# ============================================================
# EXPORTAR EXCEL
# ============================================================

def exportar_reporte_mensual_excel(request):

    hoy = timezone.localdate()
    ahora = timezone.localtime()

    curso = request.GET.get(
        "curso",
        "1° BTI"
    )

    materia_id = request.GET.get(
        "materia"
    )

    try:

        mes = int(
            request.GET.get(
                "mes",
                hoy.month
            )
        )

    except (
        TypeError,
        ValueError
    ):

        mes = hoy.month

    try:

        anio = int(
            request.GET.get(
                "anio",
                hoy.year
            )
        )

    except (
        TypeError,
        ValueError
    ):

        anio = hoy.year

    materia = None

    if materia_id:

        materia = (
            Materia.objects
            .filter(
                id=materia_id,
                activo=True
            )
            .first()
        )

    if materia is None:

        materia = (
            Materia.objects
            .filter(
                activo=True
            )
            .order_by("nombre")
            .first()
        )

    if not materia:

        return HttpResponse(
            "No existe una materia activa.",
            status=400
        )

    primer_dia = date(
        anio,
        mes,
        1
    )

    ultimo_dia = date(
        anio,
        mes,
        calendar.monthrange(
            anio,
            mes
        )[1]
    )

    horarios = (
        HorarioClase.objects
        .filter(
            anio_lectivo=anio,
            curso=curso,
            materia=materia,
            activo=True
        )
        .order_by(
            "dia_semana",
            "hora_inicio"
        )
    )

    sesiones = []

    fecha_actual = primer_dia

    while fecha_actual <= ultimo_dia:

        for horario in horarios.filter(
            dia_semana=(
                fecha_actual.weekday()
            )
        ):

            if fecha_actual < hoy:

                finalizada = True

            elif fecha_actual > hoy:

                finalizada = False

            else:

                finalizada = (
                    ahora.time()
                    >= horario.hora_fin
                )

            sesiones.append(
                {
                    "fecha": fecha_actual,
                    "horario": horario,
                    "finalizada": (
                        finalizada
                    ),
                }
            )

        fecha_actual += timedelta(
            days=1
        )

    alumnos = (
        Alumno.objects
        .filter(
            activo=True,
            curso=curso
        )
        .order_by(
            "apellidos",
            "nombres"
        )
    )

    asistencias = (
        Asistencia.objects
        .filter(
            alumno__curso=curso,
            materia=materia,
            fecha__range=[
                primer_dia,
                ultimo_dia
            ]
        )
        .select_related("alumno")
    )

    mapa = {}

    for asistencia in asistencias:

        clave = (
            asistencia.alumno_id,
            asistencia.fecha
        )

        if clave not in mapa:

            mapa[clave] = asistencia

    wb = Workbook()

    ws = wb.active
    ws.title = "Asistencia"

    azul = "0B3D91"
    verde = "C6EFCE"
    rojo = "FFC7CE"
    amarillo = "FFF2CC"

    lado = Side(
        style="thin",
        color="808080"
    )

    borde = Border(
        left=lado,
        right=lado,
        top=lado,
        bottom=lado,
    )

    total_columnas = (
        3
        + len(sesiones)
        + 3
    )

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=total_columnas,
    )

    ws["A1"] = (
        "COLEGIO NACIONAL "
        "GRAL. JOSÉ ELIZARDO AQUINO - LUQUE"
    )

    ws["A1"].font = Font(
        bold=True,
        size=14,
        color="FFFFFF"
    )

    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=azul
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=total_columnas,
    )

    ws["A2"] = (
        "PLANILLA MENSUAL DE ASISTENCIA"
    )

    ws["A2"].font = Font(
        bold=True,
        size=16
    )

    ws["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    meses = {
        1: "ENERO",
        2: "FEBRERO",
        3: "MARZO",
        4: "ABRIL",
        5: "MAYO",
        6: "JUNIO",
        7: "JULIO",
        8: "AGOSTO",
        9: "SEPTIEMBRE",
        10: "OCTUBRE",
        11: "NOVIEMBRE",
        12: "DICIEMBRE",
    }

    ws["A4"] = "CURSO:"
    ws["B4"] = curso

    ws["D4"] = "MATERIA:"
    ws["E4"] = materia.nombre

    ws["G4"] = "MES:"
    ws["H4"] = meses.get(
        mes,
        ""
    )

    ws["J4"] = "AÑO:"
    ws["K4"] = anio

    fila_header = 6

    encabezados = [
        "N°",
        "CÉDULA",
        "ALUMNO",
    ]

    for sesion in sesiones:

        encabezados.append(
            (
                sesion["fecha"]
                .strftime("%d/%m")
            )
        )

    encabezados.extend(
        [
            "P",
            "A",
            "%",
        ]
    )

    for columna, texto in enumerate(
        encabezados,
        start=1
    ):

        celda = ws.cell(
            fila_header,
            columna,
            texto
        )

        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=azul
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        celda.border = borde

    fila_excel = (
        fila_header
        + 1
    )

    for numero, alumno in enumerate(
        alumnos,
        start=1
    ):

        ws.cell(
            fila_excel,
            1,
            numero
        )

        ws.cell(
            fila_excel,
            2,
            alumno.cedula
        )

        ws.cell(
            fila_excel,
            3,
            (
                f"{alumno.apellidos}, "
                f"{alumno.nombres}"
            )
        )

        presentes = 0
        ausentes = 0

        columna = 4

        for sesion in sesiones:

            asistencia = mapa.get(
                (
                    alumno.id,
                    sesion["fecha"]
                )
            )

            celda = ws.cell(
                fila_excel,
                columna
            )

            if asistencia:

                celda.value = (
                    asistencia.hora
                    .strftime("%H:%M")
                    if asistencia.hora
                    else "P"
                )

                celda.fill = PatternFill(
                    "solid",
                    fgColor=verde
                )

                presentes += 1

            elif sesion["finalizada"]:

                celda.value = "A"

                celda.fill = PatternFill(
                    "solid",
                    fgColor=rojo
                )

                ausentes += 1

            else:

                celda.value = "-"

                celda.fill = PatternFill(
                    "solid",
                    fgColor=amarillo
                )

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            celda.border = borde

            columna += 1

        realizadas = (
            presentes
            + ausentes
        )

        porcentaje = (
            presentes
            / realizadas
            if realizadas
            else 0
        )

        ws.cell(
            fila_excel,
            columna,
            presentes
        )

        ws.cell(
            fila_excel,
            columna + 1,
            ausentes
        )

        ws.cell(
            fila_excel,
            columna + 2,
            porcentaje
        ).number_format = "0%"

        for c in range(
            1,
            columna + 3
        ):

            celda = ws.cell(
                fila_excel,
                c
            )

            celda.border = borde

            celda.alignment = Alignment(
                horizontal=(
                    "left"
                    if c == 3
                    else "center"
                ),
                vertical="center"
            )

        fila_excel += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 42

    for c in range(
        4,
        4 + len(sesiones)
    ):

        ws.column_dimensions[
            get_column_letter(c)
        ].width = 11

    ws.freeze_panes = "D7"

    ws.page_setup.orientation = (
        "landscape"
    )

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    nombre_archivo = (
        f"asistencia_"
        f"{curso}_"
        f"{materia.nombre}_"
        f"{mes:02d}_"
        f"{anio}.xlsx"
    )

    nombre_archivo = (
        nombre_archivo
        .replace("°", "")
        .replace(" ", "_")
        .replace("/", "_")
    )

    response = HttpResponse(
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; '
        f'filename="{nombre_archivo}"'
    )

    wb.save(response)

    return response


# ============================================================
# EXPORTAR REPORTE MENSUAL A PDF
# VERSIÓN FINAL
#
# - Hoja Oficio 8,5 x 13 pulgadas
# - Orientación horizontal
# - Márgenes de 2 cm
# - Logo del colegio
# - 28 alumnos en una sola hoja
# - Nombres alineados a la izquierda
# - Todos los textos centrados verticalmente
# - N°, C.I., fechas, registros, P, A y % centrados
# ============================================================

def exportar_reporte_mensual_pdf(request):

    hoy = timezone.localdate()
    ahora = timezone.localtime()

    # ========================================================
    # PARÁMETROS
    # ========================================================

    curso = request.GET.get(
        "curso",
        "1° BTI"
    )

    materia_id = request.GET.get(
        "materia"
    )

    try:
        mes = int(
            request.GET.get(
                "mes",
                hoy.month
            )
        )
    except (TypeError, ValueError):
        mes = hoy.month

    if mes < 1 or mes > 12:
        mes = hoy.month

    try:
        anio = int(
            request.GET.get(
                "anio",
                hoy.year
            )
        )
    except (TypeError, ValueError):
        anio = hoy.year

    # ========================================================
    # MATERIA
    # ========================================================

    materia = None

    if materia_id:

        materia = (
            Materia.objects
            .filter(
                id=materia_id,
                activo=True
            )
            .first()
        )

    if materia is None:

        materia = (
            Materia.objects
            .filter(
                activo=True
            )
            .order_by("nombre")
            .first()
        )

    if not materia:

        return HttpResponse(
            "No existe una materia activa.",
            status=400
        )

    # ========================================================
    # RANGO DEL MES
    # ========================================================

    primer_dia = date(
        anio,
        mes,
        1
    )

    ultimo_dia = date(
        anio,
        mes,
        calendar.monthrange(
            anio,
            mes
        )[1]
    )

    # ========================================================
    # HORARIOS DE LA MATERIA
    # ========================================================

    horarios = (
        HorarioClase.objects
        .filter(
            anio_lectivo=anio,
            curso=curso,
            materia=materia,
            activo=True
        )
        .order_by(
            "dia_semana",
            "hora_inicio"
        )
    )

    # ========================================================
    # GENERAR LAS CLASES REALES DEL MES
    # ========================================================

    sesiones = []

    fecha_actual = primer_dia

    while fecha_actual <= ultimo_dia:

        horarios_dia = horarios.filter(
            dia_semana=fecha_actual.weekday()
        )

        for horario in horarios_dia:

            if fecha_actual < hoy:

                finalizada = True

            elif fecha_actual > hoy:

                finalizada = False

            else:

                finalizada = (
                    ahora.time()
                    >= horario.hora_fin
                )

            sesiones.append(
                {
                    "fecha": fecha_actual,
                    "horario": horario,
                    "finalizada": finalizada,
                }
            )

        fecha_actual += timedelta(days=1)

    # ========================================================
    # ALUMNOS
    # ========================================================

    alumnos = (
        Alumno.objects
        .filter(
            activo=True,
            curso=curso
        )
        .order_by(
            "apellidos",
            "nombres"
        )
    )

    # ========================================================
    # ASISTENCIAS
    # ========================================================

    asistencias = (
        Asistencia.objects
        .filter(
            alumno__curso=curso,
            materia=materia,
            fecha__range=[
                primer_dia,
                ultimo_dia
            ]
        )
        .select_related(
            "alumno",
            "materia"
        )
        .order_by(
            "fecha",
            "hora"
        )
    )

    mapa_asistencias = {}

    for asistencia in asistencias:

        clave = (
            asistencia.alumno_id,
            asistencia.fecha
        )

        if clave not in mapa_asistencias:

            mapa_asistencias[
                clave
            ] = asistencia

    # ========================================================
    # CREAR BUFFER DEL PDF
    # ========================================================

    buffer = BytesIO()

    # ========================================================
    # HOJA OFICIO
    # 8,5 x 13 pulgadas
    # ========================================================

    OFICIO = (
        8.5 * inch,
        13 * inch
    )

    # ========================================================
    # DOCUMENTO
    # ========================================================

    documento = SimpleDocTemplate(
        buffer,

        pagesize=landscape(
            OFICIO
        ),

        # ----------------------------------------------------
        # MÁRGENES 2 CM
        # ----------------------------------------------------

        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,

        title=(
            "Planilla mensual de asistencia"
        ),

        author=(
            "Sistema de Asistencia Escolar"
        ),
    )

    elementos = []

    estilos = getSampleStyleSheet()

    # ========================================================
    # ESTILO COLEGIO
    # ========================================================

    estilo_colegio = ParagraphStyle(
        "ColegioPDF",

        parent=estilos["Normal"],

        fontName="Helvetica-Bold",

        fontSize=8.5,

        leading=9,

        alignment=TA_CENTER,

        spaceBefore=0,

        spaceAfter=0,
    )

    # ========================================================
    # ESTILO TÍTULO
    # ========================================================

    estilo_titulo = ParagraphStyle(
        "TituloPDF",

        parent=estilos["Normal"],

        fontName="Helvetica-Bold",

        fontSize=11,

        leading=12,

        alignment=TA_CENTER,

        spaceBefore=1,

        spaceAfter=4,
    )

    # ========================================================
    # ESTILO DATOS GENERALES
    # ========================================================

    estilo_datos = ParagraphStyle(
        "DatosPDF",

        parent=estilos["Normal"],

        fontName="Helvetica",

        fontSize=7,

        leading=7,

        alignment=TA_CENTER,

        spaceBefore=0,

        spaceAfter=0,
    )

    # ========================================================
    # ESTILO NOMBRE DEL ALUMNO
    # ========================================================

    estilo_nombre = ParagraphStyle(
        "NombreAlumnoPDF",

        parent=estilos["Normal"],

        fontName="Helvetica",

        fontSize=5.5,

        leading=5.8,

        alignment=TA_LEFT,

        spaceBefore=0,

        spaceAfter=0,

        leftIndent=0,

        rightIndent=0,
    )

    # ========================================================
    # ESTILO CENTRADO PARA:
    #
    # N°
    # CÉDULA
    # HORAS
    # A
    # P
    # %
    # ========================================================

    estilo_celda_centro = ParagraphStyle(
        "CeldaCentroPDF",

        parent=estilos["Normal"],

        fontName="Helvetica",

        fontSize=5.5,

        leading=5.8,

        alignment=TA_CENTER,

        spaceBefore=0,

        spaceAfter=0,

        leftIndent=0,

        rightIndent=0,
    )

    # ========================================================
    # ESTILO CABECERA
    # ========================================================

    estilo_cabecera = ParagraphStyle(
        "CabeceraPDF",

        parent=estilos["Normal"],

        fontName="Helvetica-Bold",

        fontSize=5.5,

        leading=5.8,

        alignment=TA_CENTER,

        textColor=colors.white,

        spaceBefore=0,

        spaceAfter=0,

        leftIndent=0,

        rightIndent=0,
    )

    # ========================================================
    # ESTILO LEYENDA
    # ========================================================

    estilo_leyenda = ParagraphStyle(
        "LeyendaPDF",

        parent=estilos["Normal"],

        fontName="Helvetica",

        fontSize=6,

        leading=7,

        spaceBefore=1,

        spaceAfter=0,
    )

    # ========================================================
    # LOGO
    # ========================================================

    ruta_logo = os.path.join(
        settings.BASE_DIR,
        "asistencia",
        "static",
        "asistencia",
        "img",
        "logo_colegio.png"
    )

    if os.path.exists(ruta_logo):

        logo = Image(
            ruta_logo,

            width=1.20 * cm,

            height=1.20 * cm
        )

        logo.hAlign = "CENTER"

        elementos.append(
            logo
        )

        elementos.append(
            Spacer(
                1,
                0.03 * cm
            )
        )

    # ========================================================
    # NOMBRE DEL COLEGIO
    # ========================================================

    elementos.append(
        Paragraph(
            "COLEGIO NACIONAL",
            estilo_colegio
        )
    )

    elementos.append(
        Paragraph(
            (
                "GRAL. JOSÉ ELIZARDO "
                "AQUINO - LUQUE"
            ),
            estilo_colegio
        )
    )

    # ========================================================
    # TÍTULO
    # ========================================================

    elementos.append(
        Paragraph(
            (
                "PLANILLA MENSUAL "
                "DE ASISTENCIA"
            ),
            estilo_titulo
        )
    )

    # ========================================================
    # NOMBRES DE MESES
    # ========================================================

    meses = {
        1: "ENERO",
        2: "FEBRERO",
        3: "MARZO",
        4: "ABRIL",
        5: "MAYO",
        6: "JUNIO",
        7: "JULIO",
        8: "AGOSTO",
        9: "SEPTIEMBRE",
        10: "OCTUBRE",
        11: "NOVIEMBRE",
        12: "DICIEMBRE",
    }

    nombre_mes = meses.get(
        mes,
        ""
    )

    # ========================================================
    # INFORMACIÓN GENERAL
    # ========================================================

    datos_generales = [
        [
            Paragraph(
                (
                    f"<b>CURSO:</b> "
                    f"{curso}"
                ),
                estilo_datos
            ),

            Paragraph(
                (
                    f"<b>MATERIA:</b> "
                    f"{materia.nombre}"
                ),
                estilo_datos
            ),

            Paragraph(
                (
                    f"<b>MES:</b> "
                    f"{nombre_mes}"
                ),
                estilo_datos
            ),

            Paragraph(
                (
                    f"<b>AÑO:</b> "
                    f"{anio}"
                ),
                estilo_datos
            ),
        ]
    ]

    # ========================================================
    # TABLA DE INFORMACIÓN GENERAL
    # ========================================================

    tabla_datos = Table(
        datos_generales,

        colWidths=[
            5.0 * cm,
            10.5 * cm,
            6.5 * cm,
            7.02 * cm,
        ],

        rowHeights=0.46 * cm,
    )

    tabla_datos.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, -1),
                colors.HexColor(
                    "#E9ECEF"
                )
            ),

            (
                "BOX",
                (0, 0),
                (-1, -1),
                0.45,
                colors.HexColor(
                    "#555555"
                )
            ),

            (
                "INNERGRID",
                (0, 0),
                (-1, -1),
                0.30,
                colors.HexColor(
                    "#999999"
                )
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                0
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                0
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                2
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                2
            ),
        ])
    )

    elementos.append(
        tabla_datos
    )

    elementos.append(
        Spacer(
            1,
            0.10 * cm
        )
    )

    # ========================================================
    # CABECERA PRINCIPAL
    #
    # IMPORTANTE:
    # TODO ES PARAGRAPH
    # ========================================================

    encabezado = [
        Paragraph(
            "N°",
            estilo_cabecera
        ),

        Paragraph(
            "C.I.",
            estilo_cabecera
        ),

        Paragraph(
            "ALUMNO",
            estilo_cabecera
        ),
    ]

    # ========================================================
    # FECHAS
    # ========================================================

    for sesion in sesiones:

        fecha_texto = (
            sesion["fecha"]
            .strftime("%d/%m")
        )

        encabezado.append(
            Paragraph(
                fecha_texto,
                estilo_cabecera
            )
        )

    # ========================================================
    # P A %
    # ========================================================

    encabezado.extend(
        [
            Paragraph(
                "P",
                estilo_cabecera
            ),

            Paragraph(
                "A",
                estilo_cabecera
            ),

            Paragraph(
                "%",
                estilo_cabecera
            ),
        ]
    )

    datos_tabla = [
        encabezado
    ]

    estados_celdas = []

    # ========================================================
    # ALUMNOS
    # ========================================================

    for numero, alumno in enumerate(
        alumnos,
        start=1
    ):

        # ====================================================
        # NÚMERO
        # ====================================================

        numero_pdf = Paragraph(
            str(numero),
            estilo_celda_centro
        )

        # ====================================================
        # CÉDULA
        # ====================================================

        cedula_pdf = Paragraph(
            str(
                alumno.cedula
                or ""
            ),
            estilo_celda_centro
        )

        # ====================================================
        # NOMBRE
        # ====================================================

        nombre_pdf = Paragraph(
            (
                f"{alumno.apellidos}, "
                f"{alumno.nombres}"
            ),
            estilo_nombre
        )

        fila = [
            numero_pdf,
            cedula_pdf,
            nombre_pdf,
        ]

        presentes = 0
        ausentes = 0

        estados_fila = []

        # ====================================================
        # SESIONES
        # ====================================================

        for sesion in sesiones:

            asistencia = (
                mapa_asistencias.get(
                    (
                        alumno.id,
                        sesion["fecha"]
                    )
                )
            )

            # ------------------------------------------------
            # PRESENTE
            # ------------------------------------------------

            if asistencia:

                if asistencia.hora:

                    valor = (
                        asistencia.hora
                        .strftime("%H:%M")
                    )

                else:

                    valor = "P"

                presentes += 1

                estado = "PRESENTE"

            # ------------------------------------------------
            # AUSENTE
            # ------------------------------------------------

            elif sesion["finalizada"]:

                valor = "A"

                ausentes += 1

                estado = "AUSENTE"

            # ------------------------------------------------
            # PENDIENTE
            # ------------------------------------------------

            else:

                valor = "-"

                estado = "PENDIENTE"

            # =================================================
            # IMPORTANTE:
            #
            # HORA / A / -
            # TAMBIÉN SON PARAGRAPH
            # =================================================

            fila.append(
                Paragraph(
                    str(valor),
                    estilo_celda_centro
                )
            )

            estados_fila.append(
                estado
            )

        # ====================================================
        # PORCENTAJE
        # ====================================================

        clases_realizadas = (
            presentes
            + ausentes
        )

        if clases_realizadas:

            porcentaje = round(
                presentes
                * 100
                / clases_realizadas
            )

        else:

            porcentaje = 0

        # ====================================================
        # P / A / %
        # TAMBIÉN COMO PARAGRAPH
        # ====================================================

        fila.extend(
            [
                Paragraph(
                    str(presentes),
                    estilo_celda_centro
                ),

                Paragraph(
                    str(ausentes),
                    estilo_celda_centro
                ),

                Paragraph(
                    f"{porcentaje}%",
                    estilo_celda_centro
                ),
            ]
        )

        datos_tabla.append(
            fila
        )

        estados_celdas.append(
            estados_fila
        )

    # ========================================================
    # ANCHO DISPONIBLE
    #
    # OFICIO HORIZONTAL:
    #
    # 13 pulgadas = 33,02 cm
    #
    # 33,02 - 2 - 2
    # = 29,02 cm
    # ========================================================

    ancho_util = 29.02

    cantidad_sesiones = len(
        sesiones
    )

    # ========================================================
    # ANCHOS FIJOS
    # ========================================================

    ancho_numero = 0.70

    ancho_cedula = 2.10

    ancho_alumno = 8.20

    ancho_presente = 0.75

    ancho_ausente = 0.75

    ancho_porcentaje = 1.00

    ancho_fijo = (
        ancho_numero
        + ancho_cedula
        + ancho_alumno
        + ancho_presente
        + ancho_ausente
        + ancho_porcentaje
    )

    # ========================================================
    # ANCHO DE CADA FECHA
    # ========================================================

    if cantidad_sesiones > 0:

        ancho_fecha = (
            ancho_util
            - ancho_fijo
        ) / cantidad_sesiones

    else:

        ancho_fecha = 0

    # ========================================================
    # LISTA DE ANCHOS
    # ========================================================

    anchos = [
        ancho_numero * cm,
        ancho_cedula * cm,
        ancho_alumno * cm,
    ]

    for _ in sesiones:

        anchos.append(
            ancho_fecha * cm
        )

    anchos.extend(
        [
            ancho_presente * cm,
            ancho_ausente * cm,
            ancho_porcentaje * cm,
        ]
    )

    # ========================================================
    # ALTURA DE LAS FILAS
    #
    # 0.46 cm permite:
    # - mejor centrado
    # - 28 alumnos
    # - una sola hoja
    # ========================================================

    altura_fila = 0.46 * cm

    # ========================================================
    # TABLA PRINCIPAL
    # ========================================================

    tabla = Table(
        datos_tabla,

        colWidths=anchos,

        rowHeights=altura_fila,

        repeatRows=1,
    )

    # ========================================================
    # ESTILO GENERAL
    # ========================================================

    estilo_tabla = TableStyle([
        # ====================================================
        # CABECERA
        # ====================================================

        (
            "BACKGROUND",
            (0, 0),
            (-1, 0),
            colors.HexColor(
                "#0B3D91"
            )
        ),

        (
            "TEXTCOLOR",
            (0, 0),
            (-1, 0),
            colors.white
        ),

        # ====================================================
        # CENTRADO VERTICAL GENERAL
        #
        # ESTA REGLA AFECTA:
        # N°
        # C.I.
        # ALUMNO
        # FECHAS
        # HORAS
        # A
        # -
        # P
        # %
        # ====================================================

        (
            "VALIGN",
            (0, 0),
            (-1, -1),
            "MIDDLE"
        ),

        # ====================================================
        # CENTRADO HORIZONTAL GENERAL
        # ====================================================

        (
            "ALIGN",
            (0, 0),
            (-1, -1),
            "CENTER"
        ),

        # ====================================================
        # NOMBRE A LA IZQUIERDA
        # ====================================================

        (
            "ALIGN",
            (2, 1),
            (2, -1),
            "LEFT"
        ),

        # ====================================================
        # NOMBRE CENTRADO VERTICAL
        # ====================================================

        (
            "VALIGN",
            (2, 1),
            (2, -1),
            "MIDDLE"
        ),

        # ====================================================
        # BORDES INTERNOS
        # ====================================================

        (
            "GRID",
            (0, 0),
            (-1, -1),
            0.35,
            colors.HexColor(
                "#777777"
            )
        ),

        # ====================================================
        # BORDE EXTERIOR
        # ====================================================

        (
            "BOX",
            (0, 0),
            (-1, -1),
            0.65,
            colors.HexColor(
                "#444444"
            )
        ),

        # ====================================================
        # PADDING VERTICAL
        #
        # 0 arriba y abajo.
        # Paragraph + VALIGN MIDDLE
        # realiza el centrado.
        # ====================================================

        (
            "TOPPADDING",
            (0, 0),
            (-1, -1),
            0
        ),

        (
            "BOTTOMPADDING",
            (0, 0),
            (-1, -1),
            0
        ),

        # ====================================================
        # PADDING HORIZONTAL
        # ====================================================

        (
            "LEFTPADDING",
            (0, 0),
            (-1, -1),
            1
        ),

        (
            "RIGHTPADDING",
            (0, 0),
            (-1, -1),
            1
        ),
    ])

    # ========================================================
    # COLORES DE ASISTENCIA
    # ========================================================

    columna_inicio_sesiones = 3

    for indice_fila, estados in enumerate(
        estados_celdas,
        start=1
    ):

        for indice_sesion, estado in enumerate(
            estados
        ):

            columna = (
                columna_inicio_sesiones
                + indice_sesion
            )

            # =================================================
            # PRESENTE
            # =================================================

            if estado == "PRESENTE":

                estilo_tabla.add(
                    "BACKGROUND",
                    (
                        columna,
                        indice_fila
                    ),
                    (
                        columna,
                        indice_fila
                    ),
                    colors.HexColor(
                        "#C6EFCE"
                    )
                )

                estilo_tabla.add(
                    "TEXTCOLOR",
                    (
                        columna,
                        indice_fila
                    ),
                    (
                        columna,
                        indice_fila
                    ),
                    colors.HexColor(
                        "#006100"
                    )
                )

            # =================================================
            # AUSENTE
            # =================================================

            elif estado == "AUSENTE":

                estilo_tabla.add(
                    "BACKGROUND",
                    (
                        columna,
                        indice_fila
                    ),
                    (
                        columna,
                        indice_fila
                    ),
                    colors.HexColor(
                        "#FFC7CE"
                    )
                )

                estilo_tabla.add(
                    "TEXTCOLOR",
                    (
                        columna,
                        indice_fila
                    ),
                    (
                        columna,
                        indice_fila
                    ),
                    colors.HexColor(
                        "#9C0006"
                    )
                )

            # =================================================
            # PENDIENTE
            # =================================================

            else:

                estilo_tabla.add(
                    "BACKGROUND",
                    (
                        columna,
                        indice_fila
                    ),
                    (
                        columna,
                        indice_fila
                    ),
                    colors.HexColor(
                        "#FFF2CC"
                    )
                )

                estilo_tabla.add(
                    "TEXTCOLOR",
                    (
                        columna,
                        indice_fila
                    ),
                    (
                        columna,
                        indice_fila
                    ),
                    colors.HexColor(
                        "#856404"
                    )
                )

    tabla.setStyle(
        estilo_tabla
    )

    elementos.append(
        tabla
    )

    # ========================================================
    # ESPACIO
    # ========================================================

    elementos.append(
        Spacer(
            1,
            0.08 * cm
        )
    )

    # ========================================================
    # LEYENDA
    # ========================================================

    elementos.append(
        Paragraph(
            (
                "<b>Referencia:</b> "
                "Hora = Presente | "
                "A = Ausente | "
                "- = Pendiente"
            ),
            estilo_leyenda
        )
    )

    # ========================================================
    # GENERAR PDF
    # ========================================================

    documento.build(
        elementos
    )

    pdf = buffer.getvalue()

    buffer.close()

    # ========================================================
    # NOMBRE DEL ARCHIVO
    # ========================================================

    nombre_archivo = (
        f"asistencia_"
        f"{curso}_"
        f"{materia.nombre}_"
        f"{mes:02d}_"
        f"{anio}.pdf"
    )

    nombre_archivo = (
        nombre_archivo
        .replace("°", "")
        .replace(" ", "_")
        .replace("/", "_")
    )

    # ========================================================
    # RESPUESTA
    # ========================================================

    response = HttpResponse(
        pdf,
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; '
        f'filename="{nombre_archivo}"'
    )

    return response

    hoy = timezone.localdate()
    ahora = timezone.localtime()

    curso = request.GET.get(
        "curso",
        "1° BTI"
    )

    materia_id = request.GET.get(
        "materia"
    )

    try:

        mes = int(
            request.GET.get(
                "mes",
                hoy.month
            )
        )

    except (
        TypeError,
        ValueError
    ):

        mes = hoy.month

    try:

        anio = int(
            request.GET.get(
                "anio",
                hoy.year
            )
        )

    except (
        TypeError,
        ValueError
    ):

        anio = hoy.year

    # --------------------------------------------------------
    # MATERIA
    # --------------------------------------------------------

    materia = None

    if materia_id:

        materia = (
            Materia.objects
            .filter(
                id=materia_id,
                activo=True
            )
            .first()
        )

    if materia is None:

        materia = (
            Materia.objects
            .filter(
                activo=True
            )
            .order_by("nombre")
            .first()
        )

    if not materia:

        return HttpResponse(
            "No existe una materia activa.",
            status=400
        )

    # --------------------------------------------------------
    # RANGO DEL MES
    # --------------------------------------------------------

    primer_dia = date(
        anio,
        mes,
        1
    )

    ultimo_dia = date(
        anio,
        mes,
        calendar.monthrange(
            anio,
            mes
        )[1]
    )

    # --------------------------------------------------------
    # HORARIOS
    # --------------------------------------------------------

    horarios = (
        HorarioClase.objects
        .filter(
            anio_lectivo=anio,
            curso=curso,
            materia=materia,
            activo=True
        )
        .order_by(
            "dia_semana",
            "hora_inicio"
        )
    )

    sesiones = []

    fecha_actual = primer_dia

    while fecha_actual <= ultimo_dia:

        horarios_dia = (
            horarios.filter(
                dia_semana=(
                    fecha_actual.weekday()
                )
            )
        )

        for horario in horarios_dia:

            if fecha_actual < hoy:

                finalizada = True

            elif fecha_actual > hoy:

                finalizada = False

            else:

                finalizada = (
                    ahora.time()
                    >= horario.hora_fin
                )

            sesiones.append(
                {
                    "fecha": fecha_actual,
                    "horario": horario,
                    "finalizada": (
                        finalizada
                    ),
                }
            )

        fecha_actual += timedelta(
            days=1
        )

    # --------------------------------------------------------
    # ALUMNOS
    # --------------------------------------------------------

    alumnos = (
        Alumno.objects
        .filter(
            activo=True,
            curso=curso
        )
        .order_by(
            "apellidos",
            "nombres"
        )
    )

    # --------------------------------------------------------
    # ASISTENCIAS
    # --------------------------------------------------------

    asistencias = (
        Asistencia.objects
        .filter(
            alumno__curso=curso,
            materia=materia,
            fecha__range=[
                primer_dia,
                ultimo_dia
            ]
        )
        .select_related(
            "alumno",
            "materia"
        )
        .order_by(
            "fecha",
            "hora"
        )
    )

    mapa = {}

    for asistencia in asistencias:

        clave = (
            asistencia.alumno_id,
            asistencia.fecha
        )

        if clave not in mapa:

            mapa[clave] = asistencia

    # ========================================================
    # PDF
    # ========================================================

    buffer = BytesIO()

    documento = SimpleDocTemplate(
        buffer,

        pagesize=landscape(
            OFICIO
        ),

        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,

        title=(
            "Planilla mensual "
            "de asistencia"
        ),

        author=(
            "Sistema de Asistencia Escolar"
        ),
    )

    elementos = []

    estilos = getSampleStyleSheet()

    # ========================================================
    # ESTILOS
    # ========================================================

    estilo_colegio = ParagraphStyle(
        "ColegioPDF",
        parent=estilos["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=9,
        alignment=TA_CENTER,
        spaceBefore=0,
        spaceAfter=0,
    )

    estilo_titulo = ParagraphStyle(
        "TituloPDF",
        parent=estilos["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=12,
        alignment=TA_CENTER,
        spaceBefore=1,
        spaceAfter=4,
    )

    estilo_datos = ParagraphStyle(
        "DatosPDF",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=7,
        alignment=TA_CENTER,
        spaceBefore=0,
        spaceAfter=0,
    )

    # ========================================================
    # ESTILO ESPECÍFICO DEL NOMBRE DEL ALUMNO
    #
    # Esto mejora mucho el centrado visual.
    # ========================================================

    estilo_nombre = ParagraphStyle(
        "NombreAlumnoPDF",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=5.7,
        leading=6,
        alignment=TA_LEFT,
        spaceBefore=0,
        spaceAfter=0,
        leftIndent=0,
        rightIndent=0,
    )

    estilo_leyenda = ParagraphStyle(
        "LeyendaPDF",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=6,
        leading=7,
        spaceBefore=1,
        spaceAfter=0,
    )

    # ========================================================
    # LOGO
    # ========================================================

    ruta_logo = os.path.join(
        settings.BASE_DIR,
        "asistencia",
        "static",
        "asistencia",
        "img",
        "logo_colegio.png"
    )

    if os.path.exists(
        ruta_logo
    ):

        logo = Image(
            ruta_logo,
            width=1.20 * cm,
            height=1.20 * cm
        )

        logo.hAlign = "CENTER"

        elementos.append(logo)

        elementos.append(
            Spacer(
                1,
                0.04 * cm
            )
        )

    # ========================================================
    # ENCABEZADO
    # ========================================================

    elementos.append(
        Paragraph(
            "COLEGIO NACIONAL",
            estilo_colegio
        )
    )

    elementos.append(
        Paragraph(
            (
                "GRAL. JOSÉ ELIZARDO "
                "AQUINO - LUQUE"
            ),
            estilo_colegio
        )
    )

    elementos.append(
        Paragraph(
            (
                "PLANILLA MENSUAL "
                "DE ASISTENCIA"
            ),
            estilo_titulo
        )
    )

    meses = {
        1: "ENERO",
        2: "FEBRERO",
        3: "MARZO",
        4: "ABRIL",
        5: "MAYO",
        6: "JUNIO",
        7: "JULIO",
        8: "AGOSTO",
        9: "SEPTIEMBRE",
        10: "OCTUBRE",
        11: "NOVIEMBRE",
        12: "DICIEMBRE",
    }

    nombre_mes = meses.get(
        mes,
        ""
    )

    # ========================================================
    # OFICIO HORIZONTAL
    #
    # 13 pulgadas = 33,02 cm
    # - 4 cm márgenes
    # = 29,02 cm útiles
    # ========================================================

    ancho_util = 29.02

    # ========================================================
    # DATOS GENERALES
    # ========================================================

    datos_generales = [
        [
            Paragraph(
                (
                    f"<b>CURSO:</b> "
                    f"{curso}"
                ),
                estilo_datos
            ),

            Paragraph(
                (
                    f"<b>MATERIA:</b> "
                    f"{materia.nombre}"
                ),
                estilo_datos
            ),

            Paragraph(
                (
                    f"<b>MES:</b> "
                    f"{nombre_mes}"
                ),
                estilo_datos
            ),

            Paragraph(
                (
                    f"<b>AÑO:</b> "
                    f"{anio}"
                ),
                estilo_datos
            ),
        ]
    ]

    tabla_datos = Table(
        datos_generales,

        colWidths=[
            5.0 * cm,
            10.5 * cm,
            6.5 * cm,
            7.02 * cm,
        ],

        rowHeights=0.46 * cm,
    )

    tabla_datos.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, -1),
                colors.HexColor(
                    "#E9ECEF"
                )
            ),

            (
                "BOX",
                (0, 0),
                (-1, -1),
                0.45,
                colors.HexColor(
                    "#555555"
                )
            ),

            (
                "INNERGRID",
                (0, 0),
                (-1, -1),
                0.30,
                colors.HexColor(
                    "#999999"
                )
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                0
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                0
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                2
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                2
            ),
        ])
    )

    elementos.append(
        tabla_datos
    )

    elementos.append(
        Spacer(
            1,
            0.10 * cm
        )
    )

    # ========================================================
    # CABECERA
    # ========================================================

    encabezado = [
        "N°",
        "C.I.",
        "ALUMNO",
    ]

    for sesion in sesiones:

        encabezado.append(
            sesion["fecha"]
            .strftime("%d/%m")
        )

    encabezado.extend(
        [
            "P",
            "A",
            "%",
        ]
    )

    datos_tabla = [
        encabezado
    ]

    estados_celdas = []

    # ========================================================
    # FILAS DE ALUMNOS
    # ========================================================

    for numero, alumno in enumerate(
        alumnos,
        start=1
    ):

        # ====================================================
        # IMPORTANTE:
        # El nombre se convierte en Paragraph.
        # ====================================================

        nombre_pdf = Paragraph(
            (
                f"{alumno.apellidos}, "
                f"{alumno.nombres}"
            ),
            estilo_nombre
        )

        fila = [
            numero,
            alumno.cedula or "",
            nombre_pdf,
        ]

        presentes = 0
        ausentes = 0

        estados = []

        for sesion in sesiones:

            asistencia = mapa.get(
                (
                    alumno.id,
                    sesion["fecha"]
                )
            )

            if asistencia:

                valor = (
                    asistencia.hora
                    .strftime("%H:%M")
                    if asistencia.hora
                    else "P"
                )

                presentes += 1

                estados.append(
                    "PRESENTE"
                )

            elif sesion["finalizada"]:

                valor = "A"

                ausentes += 1

                estados.append(
                    "AUSENTE"
                )

            else:

                valor = "-"

                estados.append(
                    "PENDIENTE"
                )

            fila.append(
                valor
            )

        realizadas = (
            presentes
            + ausentes
        )

        porcentaje = (
            round(
                presentes
                * 100
                / realizadas
            )
            if realizadas
            else 0
        )

        fila.extend(
            [
                presentes,
                ausentes,
                f"{porcentaje}%"
            ]
        )

        datos_tabla.append(
            fila
        )

        estados_celdas.append(
            estados
        )

    # ========================================================
    # DISTRIBUCIÓN HORIZONTAL
    # ========================================================

    cantidad_sesiones = len(
        sesiones
    )

    ancho_numero = 0.70
    ancho_cedula = 2.10

    # Más espacio para nombres
    ancho_alumno = 8.20

    ancho_presente = 0.75
    ancho_ausente = 0.75
    ancho_porcentaje = 1.00

    ancho_fijo = (
        ancho_numero
        + ancho_cedula
        + ancho_alumno
        + ancho_presente
        + ancho_ausente
        + ancho_porcentaje
    )

    if cantidad_sesiones:

        ancho_fecha = (
            ancho_util
            - ancho_fijo
        ) / cantidad_sesiones

    else:

        ancho_fecha = 0

    anchos = [
        ancho_numero * cm,
        ancho_cedula * cm,
        ancho_alumno * cm,
    ]

    for _ in sesiones:

        anchos.append(
            ancho_fecha * cm
        )

    anchos.extend(
        [
            ancho_presente * cm,
            ancho_ausente * cm,
            ancho_porcentaje * cm,
        ]
    )

    # ========================================================
    # TABLA PRINCIPAL
    # ========================================================

    tabla = Table(
        datos_tabla,

        colWidths=anchos,

        repeatRows=1,

        # ====================================================
        # MÁS ALTURA PARA CENTRADO VISUAL
        # ====================================================

        rowHeights=0.46 * cm,
    )

    # ========================================================
    # ESTILO
    # ========================================================

    estilo_tabla = TableStyle([
        # ----------------------------------------------------
        # CABECERA
        # ----------------------------------------------------

        (
            "BACKGROUND",
            (0, 0),
            (-1, 0),
            colors.HexColor(
                "#0B3D91"
            )
        ),

        (
            "TEXTCOLOR",
            (0, 0),
            (-1, 0),
            colors.white
        ),

        (
            "FONTNAME",
            (0, 0),
            (-1, 0),
            "Helvetica-Bold"
        ),

        (
            "FONTSIZE",
            (0, 0),
            (-1, 0),
            6
        ),

        # ----------------------------------------------------
        # CUERPO
        # ----------------------------------------------------

        (
            "FONTNAME",
            (0, 1),
            (-1, -1),
            "Helvetica"
        ),

        (
            "FONTSIZE",
            (0, 1),
            (-1, -1),
            5.7
        ),

        # ====================================================
        # ESTA ES LA CORRECCIÓN PRINCIPAL
        #
        # TODO EL CONTENIDO CENTRADO VERTICALMENTE
        # ====================================================

        (
            "VALIGN",
            (0, 0),
            (-1, -1),
            "MIDDLE"
        ),

        # ----------------------------------------------------
        # CABECERA CENTRADA
        # ----------------------------------------------------

        (
            "ALIGN",
            (0, 0),
            (-1, 0),
            "CENTER"
        ),

        # ----------------------------------------------------
        # N° Y C.I.
        # ----------------------------------------------------

        (
            "ALIGN",
            (0, 1),
            (1, -1),
            "CENTER"
        ),

        # ----------------------------------------------------
        # ALUMNO
        # horizontal izquierda
        # vertical centro
        # ----------------------------------------------------

        (
            "ALIGN",
            (2, 1),
            (2, -1),
            "LEFT"
        ),

        (
            "VALIGN",
            (2, 1),
            (2, -1),
            "MIDDLE"
        ),

        # ----------------------------------------------------
        # FECHAS / P / A / %
        # ----------------------------------------------------

        (
            "ALIGN",
            (3, 1),
            (-1, -1),
            "CENTER"
        ),

        (
            "VALIGN",
            (3, 1),
            (-1, -1),
            "MIDDLE"
        ),

        # ----------------------------------------------------
        # BORDES
        # ----------------------------------------------------

        (
            "GRID",
            (0, 0),
            (-1, -1),
            0.35,
            colors.HexColor(
                "#777777"
            )
        ),

        (
            "BOX",
            (0, 0),
            (-1, -1),
            0.65,
            colors.HexColor(
                "#444444"
            )
        ),

        # ====================================================
        # PADDING SIMÉTRICO
        #
        # Al tener el mismo padding arriba y abajo
        # el texto queda visualmente centrado.
        # ====================================================

        (
            "TOPPADDING",
            (0, 0),
            (-1, -1),
            1
        ),

        (
            "BOTTOMPADDING",
            (0, 0),
            (-1, -1),
            1
        ),

        (
            "LEFTPADDING",
            (0, 0),
            (-1, -1),
            2
        ),

        (
            "RIGHTPADDING",
            (0, 0),
            (-1, -1),
            2
        ),
    ])

    # ========================================================
    # COLORES
    # ========================================================

    columna_inicio = 3

    for fila_indice, estados in enumerate(
        estados_celdas,
        start=1
    ):

        for sesion_indice, estado in enumerate(
            estados
        ):

            columna = (
                columna_inicio
                + sesion_indice
            )

            if estado == "PRESENTE":

                estilo_tabla.add(
                    "BACKGROUND",
                    (
                        columna,
                        fila_indice
                    ),
                    (
                        columna,
                        fila_indice
                    ),
                    colors.HexColor(
                        "#C6EFCE"
                    )
                )

                estilo_tabla.add(
                    "TEXTCOLOR",
                    (
                        columna,
                        fila_indice
                    ),
                    (
                        columna,
                        fila_indice
                    ),
                    colors.HexColor(
                        "#006100"
                    )
                )

            elif estado == "AUSENTE":

                estilo_tabla.add(
                    "BACKGROUND",
                    (
                        columna,
                        fila_indice
                    ),
                    (
                        columna,
                        fila_indice
                    ),
                    colors.HexColor(
                        "#FFC7CE"
                    )
                )

                estilo_tabla.add(
                    "TEXTCOLOR",
                    (
                        columna,
                        fila_indice
                    ),
                    (
                        columna,
                        fila_indice
                    ),
                    colors.HexColor(
                        "#9C0006"
                    )
                )

            else:

                estilo_tabla.add(
                    "BACKGROUND",
                    (
                        columna,
                        fila_indice
                    ),
                    (
                        columna,
                        fila_indice
                    ),
                    colors.HexColor(
                        "#FFF2CC"
                    )
                )

                estilo_tabla.add(
                    "TEXTCOLOR",
                    (
                        columna,
                        fila_indice
                    ),
                    (
                        columna,
                        fila_indice
                    ),
                    colors.HexColor(
                        "#856404"
                    )
                )

    tabla.setStyle(
        estilo_tabla
    )

    elementos.append(
        tabla
    )

    elementos.append(
        Spacer(
            1,
            0.08 * cm
        )
    )

    # ========================================================
    # LEYENDA
    # ========================================================

    elementos.append(
        Paragraph(
            (
                "<b>Referencia:</b> "
                "Hora = Presente | "
                "A = Ausente | "
                "- = Pendiente"
            ),
            estilo_leyenda
        )
    )

    # ========================================================
    # CREAR PDF
    # ========================================================

    documento.build(
        elementos
    )

    pdf = buffer.getvalue()

    buffer.close()

    nombre_archivo = (
        f"asistencia_"
        f"{curso}_"
        f"{materia.nombre}_"
        f"{mes:02d}_"
        f"{anio}.pdf"
    )

    nombre_archivo = (
        nombre_archivo
        .replace("°", "")
        .replace(" ", "_")
        .replace("/", "_")
    )

    response = HttpResponse(
        pdf,
        content_type=(
            "application/pdf"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; '
        f'filename="{nombre_archivo}"'
    )

    return response